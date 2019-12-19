# -*- coding:utf-8 -*-
# @Time     :2019/8/13 22:19
# @author   :Dake
# Email     :604297158@qq.com
# File      :handle_excel.py
# @Software :PyCharm
import json

from openpyxl import load_workbook
from common.handle_dir import CASE_FILE_PATH


class Excel:
    """
    处理excel类
    """

    def __init__(self, file_name, sheet_name=None):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_cases(self):
        """
        获取excel中所有用例数据
        :return:返回一个列表（列表中为多个字典组成的测试用例数据）
        """
        wb = load_workbook(self.file_name)
        if self.sheet_name is None:
            ws = wb.active
        else:
            ws = wb[self.sheet_name]
        head_tuple = tuple(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]  # 获取表第一行信息
        data_tuple = tuple(ws.iter_rows(min_row=2, min_col=1, values_only=True))  # 获取表第二行之后的信息
        data_list = []
        for t in data_tuple:
            data_dict = dict(zip(head_tuple, t))
            data_list.append(data_dict)
        return data_list

    def get_case(self, row):
        """
        获取excel中某一行的用例数据
        :return:
        """
        data_list = self.get_cases()
        return data_list[row]

    def get_param(self, row):
        """
        整理case当中需要传参的数据转化为json格式
        :return: param
        """
        data = self.get_case(row)
        param_data = data["param"]
        param = json.dumps(param_data, ensure_ascii=False)
        return param

    def write_data(self, row, col, actual, result):
        """
        对excel中写入数据
        :return:
        """
        other_wb = load_workbook(self.file_name)
        if self.sheet_name is None:
            other_ws = other_wb.active
        else:
            other_ws = other_wb[self.sheet_name]

        if isinstance(row, int) and (2 <= row <= other_ws.max_row):
            other_ws.cell(row=row, column=col, value=actual)
            other_ws.cell(row=row, column=col + 1, value=result)
            other_wb.save(self.file_name)
        else:
            print("传入的行号有误，行号应该为大于1的整数")




if __name__ == "__main__":
    file = Excel(CASE_FILE_PATH)
    print(CASE_FILE_PATH)
    a = file.get_cases()
    print(a)
